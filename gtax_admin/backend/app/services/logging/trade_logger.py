"""
Excel Trade Logger.

Maintains an Excel log file for every trade with detailed information:
- Stock name
- Entry/Exit prices
- Strategy used
- Reason for entry
- P&L tracking
"""
import os
from datetime import datetime
from typing import Dict, Any, Optional
from pathlib import Path
from loguru import logger

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not installed. Excel logging disabled. Install with: pip install openpyxl")


class ExcelTradeLogger:
    """
    Logs all trades to an Excel file with comprehensive information.
    
    Columns tracked:
    - Trade ID
    - Date/Time
    - Symbol
    - Action (BUY/SELL)
    - Quantity
    - Entry Price
    - Exit Price
    - Stop Loss
    - Take Profit
    - Strategy
    - Entry Reason
    - Signal Confidence
    - Technical Score
    - Pattern Score
    - Sentiment Score
    - P&L ($)
    - P&L (%)
    - Trade Duration
    - Status
    """
    
    HEADERS = [
        'Trade ID',
        'Date/Time',
        'Symbol',
        'Action',
        'Quantity',
        'Entry Price',
        'Exit Price',
        'Stop Loss',
        'Take Profit',
        'Strategy',
        'Entry Reason',
        'Signal Confidence',
        'Technical Score',
        'Pattern Score',
        'Sentiment Score',
        'P&L ($)',
        'P&L (%)',
        'Trade Duration',
        'Status',
        'Notes'
    ]
    
    def __init__(self, log_dir: str = "logs/trades"):
        """
        Initialize the Excel trade logger.
        
        Args:
            log_dir: Directory to store trade log files
        """
        self.log_dir = Path(log_dir)
        self.log_dir.mkdir(parents=True, exist_ok=True)
        
        self.current_file: Optional[Path] = None
        self.workbook: Optional[Workbook] = None
        
        if OPENPYXL_AVAILABLE:
            self._ensure_log_file()
    
    def _get_log_filename(self) -> str:
        """Generate log filename based on current month."""
        return f"trade_log_{datetime.now().strftime('%Y_%m')}.xlsx"
    
    def _ensure_log_file(self):
        """Ensure the log file exists and is properly formatted."""
        if not OPENPYXL_AVAILABLE:
            return
        
        filename = self._get_log_filename()
        filepath = self.log_dir / filename
        
        if filepath.exists():
            self.workbook = load_workbook(filepath)
            self.current_file = filepath
        else:
            self._create_new_log_file(filepath)
    
    def _create_new_log_file(self, filepath: Path):
        """Create a new Excel log file with headers and formatting."""
        self.workbook = Workbook()
        ws = self.workbook.active
        ws.title = "Trades"
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        for col, header in enumerate(self.HEADERS, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Set column widths
        column_widths = {
            'A': 15,  # Trade ID
            'B': 20,  # Date/Time
            'C': 12,  # Symbol
            'D': 8,   # Action
            'E': 10,  # Quantity
            'F': 12,  # Entry Price
            'G': 12,  # Exit Price
            'H': 12,  # Stop Loss
            'I': 12,  # Take Profit
            'J': 20,  # Strategy
            'K': 50,  # Entry Reason
            'L': 12,  # Signal Confidence
            'M': 12,  # Technical Score
            'N': 12,  # Pattern Score
            'O': 12,  # Sentiment Score
            'P': 12,  # P&L ($)
            'Q': 10,  # P&L (%)
            'R': 15,  # Trade Duration
            'S': 12,  # Status
            'T': 30,  # Notes
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Create Summary sheet
        self._create_summary_sheet()
        
        # Save
        self.workbook.save(filepath)
        self.current_file = filepath
        logger.info(f"Created new trade log: {filepath}")
    
    def _create_summary_sheet(self):
        """Create a summary sheet with statistics."""
        if "Summary" in self.workbook.sheetnames:
            return
        
        ws = self.workbook.create_sheet("Summary")
        
        # Headers
        summary_headers = [
            ('A1', 'Trading Statistics'),
            ('A3', 'Total Trades:'),
            ('A4', 'Winning Trades:'),
            ('A5', 'Losing Trades:'),
            ('A6', 'Win Rate:'),
            ('A7', 'Total P&L:'),
            ('A8', 'Average P&L:'),
            ('A9', 'Best Trade:'),
            ('A10', 'Worst Trade:'),
            ('A12', 'By Strategy'),
        ]
        
        for cell, value in summary_headers:
            ws[cell] = value
            if cell in ['A1', 'A12']:
                ws[cell].font = Font(bold=True, size=14)
            else:
                ws[cell].font = Font(bold=True)
        
        # These will be calculated formulas
        ws['B3'] = '=COUNTA(Trades!A:A)-1'
        ws['B4'] = '=COUNTIF(Trades!P:P,">0")'
        ws['B5'] = '=COUNTIF(Trades!P:P,"<0")'
        ws['B6'] = '=IF(B3>0,B4/B3*100,0)'
        ws['B7'] = '=SUM(Trades!P:P)'
        ws['B8'] = '=IF(B3>0,B7/B3,0)'
        ws['B9'] = '=MAX(Trades!P:P)'
        ws['B10'] = '=MIN(Trades!P:P)'
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
    
    def log_entry(
        self,
        trade_id: str,
        symbol: str,
        action: str,
        quantity: float,
        entry_price: float,
        stop_loss: Optional[float] = None,
        take_profit: Optional[float] = None,
        strategy: str = "AI_COMPOSITE",
        entry_reason: str = "",
        signal_confidence: float = 0.0,
        technical_score: float = 0.0,
        pattern_score: float = 0.0,
        sentiment_score: float = 0.0,
        notes: str = ""
    ) -> bool:
        """
        Log a trade entry (BUY order).
        
        Returns:
            True if logged successfully, False otherwise
        """
        if not OPENPYXL_AVAILABLE:
            logger.warning("Cannot log trade: openpyxl not installed")
            return False
        
        try:
            self._ensure_log_file()
            ws = self.workbook.active
            
            # Find next empty row
            next_row = ws.max_row + 1
            
            # Prepare row data
            row_data = [
                trade_id,
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                symbol,
                action,
                quantity,
                round(entry_price, 2),
                None,  # Exit price (filled on exit)
                round(stop_loss, 2) if stop_loss else None,
                round(take_profit, 2) if take_profit else None,
                strategy,
                entry_reason,
                round(signal_confidence * 100, 1),
                round(technical_score * 100, 1),
                round(pattern_score * 100, 1),
                round(sentiment_score * 100, 1),
                None,  # P&L ($)
                None,  # P&L (%)
                None,  # Trade Duration
                'OPEN',
                notes
            ]
            
            # Write data
            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=next_row, column=col, value=value)
                
                # Style the action cell
                if col == 4:  # Action column
                    if value == 'BUY':
                        cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
            
            self.workbook.save(self.current_file)
            logger.info(f"Trade entry logged: {trade_id} - {action} {quantity} {symbol} @ {entry_price}")
            return True
            
        except Exception as e:
            logger.error(f"Failed to log trade entry: {e}")
            return False
    
    def log_exit(
        self,
        trade_id: str,
        exit_price: float,
        pnl: float,
        pnl_pct: float,
        exit_reason: str = ""
    ) -> bool:
        """
        Log a trade exit (SELL order / position closed).
        
        Updates the existing entry row with exit information.
        
        Returns:
            True if logged successfully, False otherwise
        """
        if not OPENPYXL_AVAILABLE:
            return False
        
        try:
            self._ensure_log_file()
            ws = self.workbook.active
            
            # Find the row with this trade_id
            trade_row = None
            for row in range(2, ws.max_row + 1):
                if ws.cell(row=row, column=1).value == trade_id:
                    trade_row = row
                    break
            
            if trade_row is None:
                # Trade not found, log as new exit record
                logger.warning(f"Trade {trade_id} not found, creating new exit record")
                return self._log_standalone_exit(trade_id, exit_price, pnl, pnl_pct, exit_reason)
            
            # Calculate trade duration
            entry_time_str = ws.cell(row=trade_row, column=2).value
            if entry_time_str:
                try:
                    entry_time = datetime.strptime(entry_time_str, '%Y-%m-%d %H:%M:%S')
                    duration = datetime.now() - entry_time
                    duration_str = str(duration).split('.')[0]  # Remove microseconds
                except:
                    duration_str = "N/A"
            else:
                duration_str = "N/A"
            
            # Update exit information
            ws.cell(row=trade_row, column=7, value=round(exit_price, 2))  # Exit Price
            ws.cell(row=trade_row, column=16, value=round(pnl, 2))  # P&L ($)
            ws.cell(row=trade_row, column=17, value=round(pnl_pct, 2))  # P&L (%)
            ws.cell(row=trade_row, column=18, value=duration_str)  # Duration
            ws.cell(row=trade_row, column=19, value='CLOSED')  # Status
            
            # Add exit reason to notes
            current_notes = ws.cell(row=trade_row, column=20).value or ""
            ws.cell(row=trade_row, column=20, value=f"{current_notes} Exit: {exit_reason}".strip())
            
            # Color code P&L
            pnl_cell = ws.cell(row=trade_row, column=16)
            if pnl > 0:
                pnl_cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
            elif pnl < 0:
                pnl_cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
            
            self.workbook.save(self.current_file)
            logger.info(f"Trade exit logged: {trade_id} - Exit @ {exit_price}, P&L: ${pnl:.2f} ({pnl_pct:.2f}%)")
            return True
            
        except Exception as e:
            logger.error(f"Failed to log trade exit: {e}")
            return False
    
    def _log_standalone_exit(
        self,
        trade_id: str,
        exit_price: float,
        pnl: float,
        pnl_pct: float,
        exit_reason: str = ""
    ) -> bool:
        """Log an exit without a matching entry (for legacy trades)."""
        try:
            ws = self.workbook.active
            next_row = ws.max_row + 1
            
            row_data = [
                trade_id,
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                trade_id.split('-')[1] if '-' in trade_id else 'UNKNOWN',  # Try to extract symbol
                'SELL',
                None,
                None,
                round(exit_price, 2),
                None,
                None,
                'UNKNOWN',
                exit_reason,
                None, None, None, None,
                round(pnl, 2),
                round(pnl_pct, 2),
                None,
                'CLOSED',
                f"Standalone exit: {exit_reason}"
            ]
            
            for col, value in enumerate(row_data, start=1):
                ws.cell(row=next_row, column=col, value=value)
            
            self.workbook.save(self.current_file)
            return True
            
        except Exception as e:
            logger.error(f"Failed to log standalone exit: {e}")
            return False
    
    def log_trade(
        self,
        trade_id: str,
        symbol: str,
        action: str,
        quantity: float,
        price: float,
        strategy: str = "AI_COMPOSITE",
        reason: str = "",
        signal_data: Optional[Dict[str, Any]] = None,
        pnl: Optional[float] = None,
        pnl_pct: Optional[float] = None,
        stop_loss: Optional[float] = None,
        take_profit: Optional[float] = None,
        notes: str = ""
    ) -> bool:
        """
        Unified trade logging method.
        
        Automatically determines if this is an entry or exit and logs appropriately.
        
        Args:
            trade_id: Unique trade identifier
            symbol: Stock/crypto symbol
            action: BUY or SELL
            quantity: Number of shares/units
            price: Execution price
            strategy: Trading strategy name
            reason: Reason for the trade
            signal_data: AI signal data with scores
            pnl: Profit/loss (for exits)
            pnl_pct: P&L percentage (for exits)
            stop_loss: Stop loss price
            take_profit: Take profit price
            notes: Additional notes
        
        Returns:
            True if logged successfully
        """
        if not OPENPYXL_AVAILABLE:
            logger.warning("Excel logging disabled: openpyxl not installed")
            return False
        
        # Extract scores from signal data
        signal_confidence = 0.0
        technical_score = 0.0
        pattern_score = 0.0
        sentiment_score = 0.0
        
        if signal_data:
            signal_confidence = signal_data.get('confidence', 0.0)
            technical_score = signal_data.get('technical_score', 0.0)
            pattern_score = signal_data.get('pattern_score', 0.0)
            sentiment_score = signal_data.get('sentiment_score', 0.0)
            
            if not reason:
                reason = signal_data.get('reason', '')
            if not strategy:
                strategy = signal_data.get('strategy', 'AI_COMPOSITE')
        
        if action.upper() == 'BUY':
            return self.log_entry(
                trade_id=trade_id,
                symbol=symbol,
                action='BUY',
                quantity=quantity,
                entry_price=price,
                stop_loss=stop_loss,
                take_profit=take_profit,
                strategy=strategy,
                entry_reason=reason,
                signal_confidence=signal_confidence,
                technical_score=technical_score,
                pattern_score=pattern_score,
                sentiment_score=sentiment_score,
                notes=notes
            )
        else:
            # For sells, try to update existing entry first
            if pnl is not None:
                return self.log_exit(
                    trade_id=trade_id,
                    exit_price=price,
                    pnl=pnl,
                    pnl_pct=pnl_pct or 0.0,
                    exit_reason=reason
                )
            else:
                # Log as new sell without P&L info
                return self.log_entry(
                    trade_id=trade_id,
                    symbol=symbol,
                    action='SELL',
                    quantity=quantity,
                    entry_price=price,
                    strategy=strategy,
                    entry_reason=reason,
                    signal_confidence=signal_confidence,
                    notes=notes
                )
    
    def get_open_trades(self) -> list:
        """Get all currently open trades."""
        if not OPENPYXL_AVAILABLE or not self.workbook:
            return []
        
        try:
            self._ensure_log_file()
            ws = self.workbook.active
            
            open_trades = []
            for row in range(2, ws.max_row + 1):
                status = ws.cell(row=row, column=19).value
                if status == 'OPEN':
                    open_trades.append({
                        'trade_id': ws.cell(row=row, column=1).value,
                        'symbol': ws.cell(row=row, column=3).value,
                        'action': ws.cell(row=row, column=4).value,
                        'quantity': ws.cell(row=row, column=5).value,
                        'entry_price': ws.cell(row=row, column=6).value,
                        'entry_time': ws.cell(row=row, column=2).value,
                        'stop_loss': ws.cell(row=row, column=8).value,
                        'take_profit': ws.cell(row=row, column=9).value,
                    })
            
            return open_trades
            
        except Exception as e:
            logger.error(f"Failed to get open trades: {e}")
            return []
    
    def get_trade_summary(self) -> Dict[str, Any]:
        """Get summary statistics of all trades."""
        if not OPENPYXL_AVAILABLE or not self.workbook:
            return {}
        
        try:
            self._ensure_log_file()
            ws = self.workbook.active
            
            total_trades = 0
            winning_trades = 0
            losing_trades = 0
            total_pnl = 0.0
            pnls = []
            
            for row in range(2, ws.max_row + 1):
                status = ws.cell(row=row, column=19).value
                pnl = ws.cell(row=row, column=16).value
                
                if status == 'CLOSED' and pnl is not None:
                    total_trades += 1
                    total_pnl += pnl
                    pnls.append(pnl)
                    
                    if pnl > 0:
                        winning_trades += 1
                    elif pnl < 0:
                        losing_trades += 1
            
            return {
                'total_trades': total_trades,
                'winning_trades': winning_trades,
                'losing_trades': losing_trades,
                'win_rate': round(winning_trades / total_trades * 100, 2) if total_trades > 0 else 0,
                'total_pnl': round(total_pnl, 2),
                'average_pnl': round(total_pnl / total_trades, 2) if total_trades > 0 else 0,
                'best_trade': round(max(pnls), 2) if pnls else 0,
                'worst_trade': round(min(pnls), 2) if pnls else 0,
                'log_file': str(self.current_file) if self.current_file else None
            }
            
        except Exception as e:
            logger.error(f"Failed to get trade summary: {e}")
            return {}


# Global instance
trade_logger = ExcelTradeLogger()
